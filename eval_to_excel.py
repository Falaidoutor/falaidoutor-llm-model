"""
Converts eval_results.json into an Excel spreadsheet for ESI comparison analysis.

Usage:
    python eval_to_excel.py [--input PATH] [--output PATH]

Requires: openpyxl
    pip install openpyxl
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

DEFAULT_INPUT = Path(__file__).parent / "eval_results.json"
DEFAULT_OUTPUT = Path(__file__).parent / "eval_comparison.xlsx"

ESI_LABELS = {
    1: "1 - Ressuscitacao",
    2: "2 - Emergente",
    3: "3 - Urgente",
    4: "4 - Menos urgente",
    5: "5 - Nao urgente",
}

ESI_LEVELS = ["ESI-1", "ESI-2", "ESI-3", "ESI-4", "ESI-5", "Indeterminado"]

ESI_FILLS = {
    1: PatternFill("solid", fgColor="FF4444"),
    2: PatternFill("solid", fgColor="FF9933"),
    3: PatternFill("solid", fgColor="FFE135"),
    4: PatternFill("solid", fgColor="44BB44"),
    5: PatternFill("solid", fgColor="5599FF"),
}

ESI_LEVEL_FILLS = {
    "ESI-1": PatternFill("solid", fgColor="FF4444"),
    "ESI-2": PatternFill("solid", fgColor="FF9933"),
    "ESI-3": PatternFill("solid", fgColor="FFE135"),
    "ESI-4": PatternFill("solid", fgColor="44BB44"),
    "ESI-5": PatternFill("solid", fgColor="5599FF"),
    "Indeterminado": PatternFill("solid", fgColor="CCCCCC"),
}

MATCH_FILL_YES = PatternFill("solid", fgColor="C6EFCE")
MATCH_FILL_NO = PatternFill("solid", fgColor="FFC7CE")

HEADER_FILL = PatternFill("solid", fgColor="2D2D2D")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)


def _to_int(value) -> int | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    return None


def _classification(llm: dict) -> str:
    classificacao = llm.get("classificacao")
    if classificacao in ESI_LEVELS:
        return classificacao

    nivel = _to_int(llm.get("nivel"))
    if nivel in ESI_LABELS:
        return f"ESI-{nivel}"

    return "Indeterminado"


def _count_items(value) -> int:
    return len(value) if isinstance(value, list) else 0


def _join_items(value) -> str:
    if isinstance(value, list):
        return "\n".join(str(item) for item in value)
    return "" if value is None else str(value)


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 70)


def style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_detail_sheet(wb, records: list[dict]):
    ws = wb.active
    ws.title = "Resultados"

    headers = [
        "edstay_id",
        "acuity_ground_truth",
        "esi_label",
        "llm_classificacao",
        "llm_nivel",
        "llm_nome_nivel",
        "ponto_decisao",
        "recursos_estimados",
        "recursos_detalhados",
        "sinais_vitais_zona_perigo",
        "populacao_especial",
        "over_triage_aplicado",
        "llm_confianca",
        "match",
        "validation_errors",
        "validation_warnings",
        "justificativa",
        "alertas",
        "error",
    ]
    ws.append(headers)
    style_header_row(ws)
    ws.row_dimensions[1].height = 30

    for rec in records:
        esi = _to_int(rec.get("acuity_ground_truth"))
        llm = rec.get("llm_response") or {}
        classificacao = _classification(llm)
        llm_nivel = _to_int(llm.get("nivel"))
        is_match = llm_nivel == esi

        row_data = [
            rec.get("edstay_id"),
            esi,
            ESI_LABELS.get(esi, str(esi or "")),
            classificacao,
            llm_nivel,
            llm.get("nome_nivel", ""),
            llm.get("ponto_decisao_ativado", ""),
            _to_int(llm.get("recursos_estimados")),
            _join_items(llm.get("recursos_detalhados")),
            llm.get("sinais_vitais_zona_perigo", ""),
            llm.get("populacao_especial", ""),
            llm.get("over_triage_aplicado", ""),
            llm.get("confianca", ""),
            "SIM" if is_match else "NAO",
            _count_items(llm.get("validation_errors")),
            _count_items(llm.get("validation_warnings")),
            llm.get("justificativa", ""),
            _join_items(llm.get("alertas")),
            rec.get("error", ""),
        ]
        ws.append(row_data)

        row_idx = ws.max_row
        if esi in ESI_FILLS:
            ws.cell(row_idx, 2).fill = ESI_FILLS[esi]
        if classificacao in ESI_LEVEL_FILLS:
            ws.cell(row_idx, 4).fill = ESI_LEVEL_FILLS[classificacao]
        if llm_nivel in ESI_FILLS:
            ws.cell(row_idx, 5).fill = ESI_FILLS[llm_nivel]
        ws.cell(row_idx, 14).fill = MATCH_FILL_YES if is_match else MATCH_FILL_NO
        ws.cell(row_idx, 14).font = BOLD
        ws.cell(row_idx, 14).alignment = Alignment(horizontal="center")

        for col_idx in (9, 17, 18, 19):
            ws.cell(row_idx, col_idx).alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    auto_width(ws)


def build_summary_sheet(wb, records: list[dict]):
    ws = wb.create_sheet("Resumo")

    ws.append(["ACERTOS POR NIVEL ESI"])
    ws["A1"].font = Font(bold=True, size=12)

    ws.append(["ESI", "Label", "Total", "Acertos", "Taxa de Acerto (%)"])
    style_header_row(ws, row=2)

    esi_totals = {i: 0 for i in range(1, 6)}
    esi_matches = {i: 0 for i in range(1, 6)}

    for rec in records:
        esi = _to_int(rec.get("acuity_ground_truth"))
        llm_nivel = _to_int((rec.get("llm_response") or {}).get("nivel"))
        if esi in esi_totals:
            esi_totals[esi] += 1
            if llm_nivel == esi:
                esi_matches[esi] += 1

    for esi in range(1, 6):
        total = esi_totals[esi]
        matches = esi_matches[esi]
        rate = round(matches / total * 100, 1) if total > 0 else 0
        ws.append([esi, ESI_LABELS[esi], total, matches, rate])
        row_idx = ws.max_row
        ws.cell(row_idx, 1).fill = ESI_FILLS[esi]
        ws.cell(row_idx, 1).font = BOLD

    grand_total = sum(esi_totals.values())
    grand_matches = sum(esi_matches.values())
    grand_rate = round(grand_matches / grand_total * 100, 1) if grand_total > 0 else 0
    ws.append(["TOTAL", "", grand_total, grand_matches, grand_rate])
    for col in range(1, 6):
        ws.cell(ws.max_row, col).font = BOLD
        ws.cell(ws.max_row, col).fill = PatternFill("solid", fgColor="DDDDDD")

    ws.append([])

    matrix_start_row = ws.max_row + 1
    ws.cell(matrix_start_row, 1).value = "MATRIZ DE CONFUSAO (ESI Ground Truth vs. ESI LLM)"
    ws.cell(matrix_start_row, 1).font = Font(bold=True, size=12)

    header_row = matrix_start_row + 1
    ws.cell(header_row, 1).value = "GT \\ LLM ->"
    ws.cell(header_row, 1).fill = HEADER_FILL
    ws.cell(header_row, 1).font = HEADER_FONT

    for col_idx, level in enumerate(ESI_LEVELS, start=2):
        cell = ws.cell(header_row, col_idx)
        cell.value = level
        cell.fill = ESI_LEVEL_FILLS.get(level, PatternFill("solid", fgColor="CCCCCC"))
        cell.font = BOLD
        cell.alignment = Alignment(horizontal="center")

    matrix = {esi: {lvl: 0 for lvl in ESI_LEVELS} for esi in range(1, 6)}
    for rec in records:
        esi = _to_int(rec.get("acuity_ground_truth"))
        classificacao = _classification(rec.get("llm_response") or {})
        if esi in matrix:
            matrix[esi][classificacao] += 1

    for esi in range(1, 6):
        row_idx = header_row + esi
        label_cell = ws.cell(row_idx, 1)
        label_cell.value = ESI_LABELS[esi]
        label_cell.fill = ESI_FILLS[esi]
        label_cell.font = BOLD

        expected_level = f"ESI-{esi}"
        for col_idx, level in enumerate(ESI_LEVELS, start=2):
            count = matrix[esi][level]
            cell = ws.cell(row_idx, col_idx)
            cell.value = count if count > 0 else ""
            cell.alignment = Alignment(horizontal="center")
            if level == expected_level and count > 0:
                cell.fill = MATCH_FILL_YES
                cell.font = Font(bold=True, color="375623")
            elif count > 0:
                cell.fill = MATCH_FILL_NO

    ws.append([])

    dist_start = ws.max_row + 1
    ws.cell(dist_start, 1).value = "DISTRIBUICAO DE CLASSIFICACOES LLM POR NIVEL ESI"
    ws.cell(dist_start, 1).font = Font(bold=True, size=12)

    dist_header = dist_start + 1
    ws.cell(dist_header, 1).value = "ESI"
    ws.cell(dist_header, 1).font = HEADER_FONT
    ws.cell(dist_header, 1).fill = HEADER_FILL
    for col_idx, level in enumerate(ESI_LEVELS, start=2):
        cell = ws.cell(dist_header, col_idx)
        cell.value = level
        cell.fill = ESI_LEVEL_FILLS.get(level, PatternFill("solid", fgColor="CCCCCC"))
        cell.font = BOLD
        cell.alignment = Alignment(horizontal="center")
    ws.cell(dist_header, len(ESI_LEVELS) + 2).value = "Total"
    ws.cell(dist_header, len(ESI_LEVELS) + 2).font = BOLD

    for esi in range(1, 6):
        row_idx = dist_header + esi
        ws.cell(row_idx, 1).value = esi
        ws.cell(row_idx, 1).fill = ESI_FILLS[esi]
        ws.cell(row_idx, 1).font = BOLD
        row_total = 0
        for col_idx, level in enumerate(ESI_LEVELS, start=2):
            count = matrix[esi][level]
            cell = ws.cell(row_idx, col_idx)
            cell.value = count if count > 0 else ""
            cell.alignment = Alignment(horizontal="center")
            if count > 0:
                cell.fill = ESI_LEVEL_FILLS.get(level, PatternFill("solid", fgColor="CCCCCC"))
                cell.font = Font(color="000000")
            row_total += count
        ws.cell(row_idx, len(ESI_LEVELS) + 2).value = row_total
        ws.cell(row_idx, len(ESI_LEVELS) + 2).font = BOLD

    auto_width(ws)


def main():
    parser = argparse.ArgumentParser(description="Convert eval_results.json to Excel comparison sheet.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    if not args.input.exists():
        print(f"ERROR: Input file not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    with args.input.open(encoding="utf-8") as f:
        records = json.load(f)

    usable = [r for r in records if r.get("llm_response") is not None]
    skipped = len(records) - len(usable)
    print(f"Loaded {len(records)} records ({skipped} without llm_response skipped).")

    wb = openpyxl.Workbook()
    build_detail_sheet(wb, usable)
    build_summary_sheet(wb, usable)

    wb.save(args.output)
    print(f"Saved: {args.output}")

    total = len(usable)
    matches = sum(
        1
        for r in usable
        if _to_int((r.get("llm_response") or {}).get("nivel")) == _to_int(r.get("acuity_ground_truth"))
    )
    accuracy = matches / total * 100 if total else 0
    print(f"\nOverall accuracy: {matches}/{total} = {accuracy:.1f}%")
    print("\nPer ESI level:")
    for esi in range(1, 6):
        group = [r for r in usable if _to_int(r.get("acuity_ground_truth")) == esi]
        hits = sum(1 for r in group if _to_int((r.get("llm_response") or {}).get("nivel")) == esi)
        rate = hits / len(group) * 100 if group else 0
        print(f"  ESI {esi} ({ESI_LABELS[esi]:>20}): {hits:>3}/{len(group):<3} = {rate:.1f}%")


if __name__ == "__main__":
    main()
